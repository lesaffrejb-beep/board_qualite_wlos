#!/usr/bin/env python3
"""
Analyse complète des audits qualité Slow Village
Expert: Marie-Claude Fortier - 15 ans d'expérience audit qualité hôtellerie plein air
"""

import openpyxl
import json
from pathlib import Path
from collections import defaultdict
from typing import Dict, List, Any, Tuple

# Configuration
AUDIT_DIR = Path("/Users/jb/Downloads/SYNTHESE_AUDIT_QUALITE")

# Mapping des codes thématiques (basé sur les données observées)
THEME_MAPPING = {
    'AFF': 'Affichage & Communication',
    'EXP': 'Expérience Client',
    'IMA': 'Image & Esthétique',
    'QUA': 'Qualité & Propreté',
    'RES': 'Restauration',
    'RH': 'Ressources Humaines',
    'SEC': 'Sécurité & Réglementation',
    'SLO': 'Slow Village (Identité Marque)'
}

def detect_sheets_structure(filepath: Path) -> List[str]:
    """Détecte les onglets du fichier Excel"""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets

def parse_detail_sheet(ws) -> Dict[str, Any]:
    """Parse l'onglet détaillé des critères"""
    data = {
        'themes': defaultdict(lambda: {
            'sous_themes': defaultdict(lambda: {
                'criteres': [],
                'stats': {'conforme': 0, 'non_conforme': 0, 'en_cours': 0, 'na': 0, 'total': 0}
            }),
            'stats': {'conforme': 0, 'non_conforme': 0, 'en_cours': 0, 'na': 0, 'total': 0}
        }),
        'global_stats': {'conforme': 0, 'non_conforme': 0, 'en_cours': 0, 'na': 0, 'total': 0}
    }
    
    current_theme = None
    current_sous_theme = None
    
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        if not any(row):
            continue
        
        # Identifier la structure (à adapter selon le format réel)
        # Généralement : Thème | Sous-thème | Critère | Statut | Remarques
        
        # Stocker les 100 premières lignes pour analyse manuelle
        if i <= 100:
            print(f"  L{i:03d}: {row[:8]}")
    
    return data

def analyze_single_file(filepath: Path) -> Dict[str, Any]:
    """Analyse complète d'un fichier d'audit"""
    site_name = filepath.stem.replace('Grille Audit ', '')
    print(f"\n{'='*80}")
    print(f"📍 SITE: {site_name}")
    print(f"{'='*80}")
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # Lister les onglets disponibles
    print(f"\n📑 Onglets disponibles: {wb.sheetnames}")
    
    # Analyser chaque onglet
    for sheet_name in wb.sheetnames:
        print(f"\n📄 Onglet: '{sheet_name}'")
        ws = wb[sheet_name]
        
        # Afficher les 100 premières lignes pour comprendre la structure
        print(f"  Dimensions: {ws.dimensions}")
        
        # Observer la structure
        if 'Détail' in sheet_name or 'Detail' in sheet_name or sheet_name == wb.sheetnames[-1]:
            print(f"\n  🔍 Analyse approfondie de '{sheet_name}':")
            parse_detail_sheet(ws)
            break  # On s'arrête après avoir trouvé l'onglet détaillé
    
    wb.close()
    
    return {
        'site': site_name,
        'sheets': wb.sheetnames if 'wb' in locals() else []
    }

def main():
    print("=" * 100)
    print("🎯 ANALYSE APPROFONDIE - AUDIT QUALITÉ GROUPE SLOW VILLAGE")
    print("=" * 100)
    
    # Analyser le premier fichier en détail pour comprendre la structure
    first_file = AUDIT_DIR / "Grille Audit Anduze.xlsx"
    
    if first_file.exists():
        result = analyze_single_file(first_file)
        print(f"\n✅ Analyse terminée pour {result['site']}")
    else:
        print(f"❌ Fichier non trouvé: {first_file}")

if __name__ == "__main__":
    main()
