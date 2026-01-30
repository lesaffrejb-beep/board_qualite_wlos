#!/usr/bin/env python3
"""
Extraction FINALE des audits Slow Village - Parser optimisé v2
Marie-Claude Fortier - Expert Audit Qualité
Données enrichies : scores pondérés, bloquants, dates, priorités
"""

import openpyxl
import json
from pathlib import Path
from collections import defaultdict
from datetime import datetime

AUDIT_DIR = Path("/Users/jb/Documents/SYNTHESE_AUDIT_QUALITE")

THEME_SHEETS = {
    'AFF': 'Affichage Obligatoire (AFF)',
    'EXP': 'Exploitation (EXP)',
    'IMA': 'Image (IMA)',
    'QUA': 'Qualité (QUA)',
    'RES': 'Résidents (RES)',
    'RH': 'RH (RH)',
    'SEC': 'Sécurité (SEC)',
    'SLO': 'Slow Impact (SLO)'
}

THEME_NAMES = {
    'AFF': 'Affichage & Communication',
    'EXP': 'Expérience Client & Exploitation',
    'IMA': 'Image & Esthétique',
    'QUA': 'Qualité & Propreté',
    'RES': 'Résidents & Propriétaires',
    'RH': 'Ressources Humaines',
    'SEC': 'Sécurité & Réglementation',
    'SLO': 'Slow Village (Identité Marque)'
}

def parse_synthese_sheet(ws):
    """Extrait les scores pondérés et métadonnées de la feuille Synthèse"""
    synthese_data = {
        'date_audit': None,
        'responsable_audit': None,
        'scores_pondérés': {},
        'total_obtenu': 0,
        'total_possible': 0,
        'taux_conformite_pondere': 0,
        'nb_bloquants': 0
    }
    
    # Extraction date et responsable (lignes 2-3)
    date_val = ws['C2'].value
    if date_val:
        if isinstance(date_val, datetime):
            synthese_data['date_audit'] = date_val.strftime('%d/%m/%Y')
        else:
            synthese_data['date_audit'] = str(date_val)
    
    resp_val = ws['C3'].value
    if resp_val:
        synthese_data['responsable_audit'] = str(resp_val)
    
    # Extraction des scores par thème (lignes 6-14)
    theme_mapping = {
        'Affichage Obligatoire': 'AFF',
        'Exploitation': 'EXP',
        'Image': 'IMA',
        'Qualité': 'QUA',
        'Résidents': 'RES',
        'RH': 'RH',
        'Sécurité': 'SEC',
        'Slow Impact': 'SLO'
    }
    
    for row in ws.iter_rows(min_row=6, max_row=14, values_only=True):
        if not row or not row[1]:
            continue
        
        theme_name = str(row[1]).strip()
        if theme_name in theme_mapping:
            theme_code = theme_mapping[theme_name]
            try:
                obtenu = float(row[2]) if row[2] else 0
                possible = float(row[3]) if row[3] else 0
                pct = float(row[4]) if row[4] else 0
                bloquants = int(row[5]) if row[5] else 0
                
                synthese_data['scores_pondérés'][theme_code] = {
                    'obtenu': obtenu,
                    'possible': possible,
                    'pourcentage': round(pct * 100, 2) if pct <= 1 else round(pct, 2),
                    'bloquants': bloquants
                }
                
                synthese_data['total_obtenu'] += obtenu
                synthese_data['total_possible'] += possible
                synthese_data['nb_bloquants'] += bloquants
            except (ValueError, TypeError):
                pass
    
    # Calcul du taux global pondéré
    if synthese_data['total_possible'] > 0:
        synthese_data['taux_conformite_pondere'] = round(
            (synthese_data['total_obtenu'] / synthese_data['total_possible']) * 100, 2
        )
    
    return synthese_data

def parse_theme_sheet(ws, theme_code):
    """Parse un onglet thématique selon la structure détectée"""
    
    theme_data = {
        'theme_code': theme_code,
        'theme_name': THEME_NAMES[theme_code],
        'criteres': [],
        'stats': {'conforme': 0, 'non_conforme': 0, 'en_cours': 0, 'na': 0, 'total': 0},
        'stats_ponderees': {'obtenu': 0, 'possible': 0},
        'sous_themes': defaultdict(lambda: {'conforme': 0, 'non_conforme': 0, 'en_cours': 0, 'na': 0, 'total': 0}),
        'priorites': {'P0': {'conforme': 0, 'non_conforme': 0, 'en_cours': 0, 'na': 0},
                      'P1': {'conforme': 0, 'non_conforme': 0, 'en_cours': 0, 'na': 0},
                      'P2': {'conforme': 0, 'non_conforme': 0, 'en_cours': 0, 'na': 0},
                      'P3': {'conforme': 0, 'non_conforme': 0, 'en_cours': 0, 'na': 0}},
        'obligatoires': {'total': 0, 'conforme': 0, 'non_conforme': 0}
    }
    
    for i, row in enumerate(ws.iter_rows(min_row=10, values_only=True), 10):
        if not any(row):
            continue
        
        # Colonnes : A=0, B=1, C=2, D=3, E=4, F=5, G=6, H=7, I=8, J=9, K=10, L=11
        # B=1: Code critère (ex: AFF-10)
        # C=2: Thème
        # D=3: LOT
        # E=4: Sous-thème
        # F=5: Critère
        # G=6: Priorité (P1, P2, P3)
        # H=7: Obligatoire (bool)
        # I=8: Conforme (bool) 
        # J=9: Non conforme (bool)
        # K=10: NA (bool)
        # L=11: Score
        
        code_critere = row[1]
        lot = row[3]
        sous_theme = row[4]
        critere_text = row[5]
        priorite = row[6]
        obligatoire = row[7] if len(row) > 7 else False
        
        # Les statuts (colonnes I à K, indices 8 à 10)
        conforme = row[8] if len(row) > 8 else False
        non_conforme = row[9] if len(row) > 9 else False
        na = row[10] if len(row) > 10 else False
        score = row[11] if len(row) > 11 else 0
        
        # Ignorer les lignes vides ou d'en-tête
        if not code_critere or not isinstance(code_critere, str):
            continue
        
        # Déterminer le statut
        if conforme:
            statut = 'conforme'
        elif non_conforme:
            statut = 'non_conforme'
        elif na:
            statut = 'na'
        else:
            statut = 'en_cours'
        
        # Normaliser priorité
        prio_norm = 'P3'
        if priorite:
            prio_str = str(priorite).upper().strip()
            if prio_str in ['P0', 'P1', 'P2', 'P3']:
                prio_norm = prio_str
        
        # Ajouter le critère
        critere_obj = {
            'code': code_critere,
            'lot': lot or 'Non catégorisé',
            'sous_theme': sous_theme or 'Général',
            'critere': critere_text or '',
            'priorite': prio_norm,
            'obligatoire': bool(obligatoire),
            'statut': statut,
            'score': score if score else 0
        }
        
        theme_data['criteres'].append(critere_obj)
        theme_data['stats'][statut] += 1
        theme_data['stats']['total'] += 1
        
        # Stats par sous-thème
        st_key = sous_theme or 'Général'
        theme_data['sous_themes'][st_key][statut] += 1
        theme_data['sous_themes'][st_key]['total'] += 1
        
        # Stats par priorité
        theme_data['priorites'][prio_norm][statut] += 1
        
        # Stats obligatoires
        if obligatoire:
            theme_data['obligatoires']['total'] += 1
            if statut == 'conforme':
                theme_data['obligatoires']['conforme'] += 1
            elif statut == 'non_conforme':
                theme_data['obligatoires']['non_conforme'] += 1
    
    # Convertir defaultdict en dict normal pour JSON
    theme_data['sous_themes'] = dict(theme_data['sous_themes'])
    
    return theme_data

def analyze_site(filepath):
    """Analyse complète d'un site"""
    site_name = filepath.stem.replace('Grille Audit ', '')
    
    print(f"\n📍 {site_name:35s}", end=' ')
    
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    
    site_data = {
        'site': site_name,
        'date_audit': None,
        'responsable_audit': None,
        'themes': {},
        'global_stats': {'conforme': 0, 'non_conforme': 0, 'en_cours': 0, 'na': 0, 'total': 0},
        'global_stats_ponderees': {'obtenu': 0, 'possible': 0, 'taux': 0},
        'priorites_globales': {'P0': {'conforme': 0, 'non_conforme': 0, 'en_cours': 0, 'na': 0},
                               'P1': {'conforme': 0, 'non_conforme': 0, 'en_cours': 0, 'na': 0},
                               'P2': {'conforme': 0, 'non_conforme': 0, 'en_cours': 0, 'na': 0},
                               'P3': {'conforme': 0, 'non_conforme': 0, 'en_cours': 0, 'na': 0}},
        'obligatoires_globaux': {'total': 0, 'conforme': 0, 'non_conforme': 0},
        'nb_bloquants': 0,
        'scores_par_theme': {}
    }
    
    # Extraction de la feuille Synthèse
    if 'Synthèse' in wb.sheetnames:
        ws_synthese = wb['Synthèse']
        synthese = parse_synthese_sheet(ws_synthese)
        site_data['date_audit'] = synthese['date_audit']
        site_data['responsable_audit'] = synthese['responsable_audit']
        site_data['global_stats_ponderees'] = {
            'obtenu': synthese['total_obtenu'],
            'possible': synthese['total_possible'],
            'taux': synthese['taux_conformite_pondere']
        }
        site_data['nb_bloquants'] = synthese['nb_bloquants']
        site_data['scores_par_theme'] = synthese['scores_pondérés']
    
    # Extraction des thèmes
    for code, sheet_name in THEME_SHEETS.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            theme_data = parse_theme_sheet(ws, code)
            site_data['themes'][code] = theme_data
            
            # Agréger stats globales
            for key in ['conforme', 'non_conforme', 'en_cours', 'na', 'total']:
                site_data['global_stats'][key] += theme_data['stats'][key]
            
            # Agréger priorités
            for prio in ['P0', 'P1', 'P2', 'P3']:
                for key in ['conforme', 'non_conforme', 'en_cours', 'na']:
                    site_data['priorites_globales'][prio][key] += theme_data['priorites'][prio][key]
            
            # Agréger obligatoires
            site_data['obligatoires_globaux']['total'] += theme_data['obligatoires']['total']
            site_data['obligatoires_globaux']['conforme'] += theme_data['obligatoires']['conforme']
            site_data['obligatoires_globaux']['non_conforme'] += theme_data['obligatoires']['non_conforme']
    
    wb.close()
    
    # Calculs dérivés
    total = site_data['global_stats']['total']
    conf = site_data['global_stats']['conforme']
    nc = site_data['global_stats']['non_conforme']
    ec = site_data['global_stats']['en_cours']
    na = site_data['global_stats']['na']
    
    taux_brut = (conf / total * 100) if total > 0 else 0
    taux_pondere = site_data['global_stats_ponderees']['taux']
    
    # Calcul écart-type (dispersion des scores par thème)
    scores_themes = [s['pourcentage'] for s in site_data['scores_par_theme'].values() if s['possible'] > 0]
    if len(scores_themes) > 1:
        moyenne = sum(scores_themes) / len(scores_themes)
        variance = sum((x - moyenne) ** 2 for x in scores_themes) / len(scores_themes)
        site_data['ecart_type_themes'] = round(variance ** 0.5, 2)
    else:
        site_data['ecart_type_themes'] = 0
    
    print(f"| {total:4d} crit | ✅ {conf:3d} | ❌ {nc:3d} | 🔄 {ec:4d} | Pondéré: {taux_pondere:5.1f}% | Bloquants: {site_data['nb_bloquants']:2d}")
    
    return site_data

def main():
    print("="*140)
    print("🎯 EXTRACTION COMPLÈTE v2 - AUDIT QUALITÉ GROUPE SLOW VILLAGE")
    print("   Données enrichies : scores pondérés | bloquants | dates | obligatoires | écarts")
    print("="*140)
    
    all_data = []
    audit_files = sorted(AUDIT_DIR.glob("Grille Audit *.xlsx"))
    
    for filepath in audit_files:
        try:
            site_data = analyze_site(filepath)
            all_data.append(site_data)
        except Exception as e:
            print(f"❌ Erreur sur {filepath.name}: {e}")
            import traceback
            traceback.print_exc()
    
    # Sauvegarde
    output_file = AUDIT_DIR / "synthese_data_complete.json"
    
    # Sauvegarde également dans data_extracted.json pour compatibilité
    simple_data = [{'site': s['site'], 'themes': {}} for s in all_data]
    simple_output = AUDIT_DIR / "data_extracted.json"
    with open(simple_output, 'w', encoding='utf-8') as f:
        json.dump(simple_data, f, ensure_ascii=False, indent=2)
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)
    
    print(f"\n{'='*140}")
    print(f"✅ {len(all_data)} sites analysés avec succès")
    print(f"💾 Données : {output_file}")
    
    # Stats globales groupe
    total_criteres = sum(s['global_stats']['total'] for s in all_data)
    total_conforme = sum(s['global_stats']['conforme'] for s in all_data)
    total_nc = sum(s['global_stats']['non_conforme'] for s in all_data)
    total_ec = sum(s['global_stats']['en_cours'] for s in all_data)
    total_na = sum(s['global_stats']['na'] for s in all_data)
    
    total_bloquants = sum(s['nb_bloquants'] for s in all_data)
    
    # Stats pondérées groupe
    total_obtenu_groupe = sum(s['global_stats_ponderees']['obtenu'] for s in all_data)
    total_possible_groupe = sum(s['global_stats_ponderees']['possible'] for s in all_data)
    taux_pondere_groupe = (total_obtenu_groupe / total_possible_groupe * 100) if total_possible_groupe > 0 else 0
    
    # Priorités groupe
    p1_nc = sum(s['priorites_globales']['P1']['non_conforme'] for s in all_data)
    p1_total = sum(sum(s['priorites_globales']['P1'].values()) for s in all_data)
    
    print(f"\n{'='*140}")
    print(f"📊 GROUPE SLOW VILLAGE - SYNTHÈSE GLOBALE")
    print(f"{'='*140}")
    print(f"   Total critères audités    : {total_criteres:,}")
    print(f"   ✅ Conforme (comptage)    : {total_conforme:,} ({total_conforme/total_criteres*100:.1f}%)")
    print(f"   ❌ Non conforme           : {total_nc:,} ({total_nc/total_criteres*100:.1f}%)")
    print(f"   🔄 En cours               : {total_ec:,} ({total_ec/total_criteres*100:.1f}%)")
    print(f"   ⚪ N/A                    : {total_na:,} ({total_na/total_criteres*100:.1f}%)")
    print(f"")
    print(f"   📊 TAUX CONFORMITÉ PONDÉRÉ : {taux_pondere_groupe:.1f}%")
    print(f"      (Total obtenu: {total_obtenu_groupe:.0f} / {total_possible_groupe:.0f})")
    print(f"")
    print(f"   🚨 Critères bloquants     : {total_bloquants}")
    print(f"   ⚠️  P1 Non conformes       : {p1_nc} sur {p1_total} critères P1 ({p1_nc/p1_total*100:.1f}%)")
    print("="*140)

if __name__ == "__main__":
    main()
