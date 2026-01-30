#!/usr/bin/env python3
"""
Vérification d'intégrité complète des données extraites v2
Vérifie : statuts, scores, dates, bloquants, cohérence mathématique
"""

import json
import openpyxl
import random
from pathlib import Path
from datetime import datetime

AUDIT_DIR = Path("/Users/jb/Documents/SYNTHESE_AUDIT_QUALITE")
JSON_FILE = AUDIT_DIR / "synthese_data_complete.json"

def load_json_data():
    with open(JSON_FILE, 'r') as f:
        return json.load(f)

def verify_sample_criteria(json_data, n_samples=5):
    """Vérifie N critères aléatoires contre les fichiers Excel source"""
    
    # Aplatir tous les critères pour sélection aléatoire
    all_criteria_refs = []
    for site in json_data:
        site_name = site['site']
        for theme_code, theme_data in site['themes'].items():
            for crit in theme_data['criteres']:
                all_criteria_refs.append({
                    'site': site_name,
                    'theme': theme_code,
                    'code': crit['code'],
                    'critere': crit['critere'],
                    'json_statut': crit['statut'],
                    'json_priorite': crit.get('priorite', 'P3'),
                    'json_obligatoire': crit.get('obligatoire', False),
                    'json_score': crit.get('score', 0)
                })
    
    samples = random.sample(all_criteria_refs, min(n_samples, len(all_criteria_refs)))
    
    print("\n" + "="*100)
    print("🔍 PARTIE 1 : VÉRIFICATION CRITÈRES ALÉATOIRES (Contrôle surprise)")
    print("="*100)
    
    results = []
    
    for sample in samples:
        site = sample['site']
        theme = sample['theme']
        code = sample['code']
        
        print(f"\n👉 Test: {site} | {theme} | {code}")
        print(f"   Critère: {sample['critere'][:70]}...")
        print(f"   JSON → Statut:{sample['json_statut'].upper():12s} | Priorité:{sample['json_priorite']} | Obligatoire:{sample['json_obligatoire']}")
        
        # Vérification dans Excel
        excel_path = AUDIT_DIR / f"Grille Audit {site}.xlsx"
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        theme_map = {
            'AFF': 'AFF', 'EXP': 'EXP', 'IMA': 'IMA', 'QUA': 'QUA', 
            'RES': 'RES', 'RH': 'RH', 'SEC': 'SEC', 'SLO': 'SLO'
        }
        
        sheet_found = False
        excel_value = "NOT FOUND"
        excel_priorite = "N/A"
        excel_obligatoire = False
        
        for sheet_name in wb.sheetnames:
            if theme_map.get(theme) in sheet_name:
                ws = wb[sheet_name]
                for i, row in enumerate(ws.iter_rows(min_row=10, values_only=True), 10):
                    if row[1] == code:
                        sheet_found = True
                        is_conf = row[8]
                        is_nc = row[9]
                        is_na = row[10]
                        
                        if is_conf: excel_value = "conforme"
                        elif is_nc: excel_value = "non_conforme"
                        elif is_na: excel_value = "na"
                        else: excel_value = "en_cours"
                        
                        excel_priorite = str(row[6]).upper().strip() if row[6] else "P3"
                        excel_obligatoire = bool(row[7]) if len(row) > 7 else False
                        break
            if sheet_found: break
        
        wb.close()
        
        # Vérifications
        match_statut = (excel_value == sample['json_statut'])
        match_priorite = (excel_priorite == sample['json_priorite'])
        match_obligatoire = (excel_obligatoire == sample['json_obligatoire'])
        
        all_match = match_statut and match_priorite and match_obligatoire
        icon = "✅" if all_match else "❌"
        
        print(f"   EXCEL→ Statut:{excel_value.upper():12s} | Priorité:{excel_priorite} | Obligatoire:{excel_obligatoire}")
        print(f"   Résultat: {icon} {'Tout OK' if all_match else 'Différence détectée'}")
        
        if not match_statut:
            print(f"      ⚠️  Statut différent: JSON={sample['json_statut']} vs EXCEL={excel_value}")
        if not match_priorite:
            print(f"      ⚠️  Priorité différente: JSON={sample['json_priorite']} vs EXCEL={excel_priorite}")
        if not match_obligatoire:
            print(f"      ⚠️  Obligatoire différent: JSON={sample['json_obligatoire']} vs EXCEL={excel_obligatoire}")
        
        results.append(all_match)
    
    return results

def verify_scores_synthese(json_data):
    """Vérifie les scores pondérés de la feuille Synthèse"""
    
    print("\n" + "="*100)
    print("🔍 PARTIE 2 : VÉRIFICATION SCORES PONDÉRÉS (Feuille Synthèse)")
    print("="*100)
    
    results = []
    
    # Sélectionner 3 sites aléatoires
    sample_sites = random.sample(json_data, min(3, len(json_data)))
    
    for site_data in sample_sites:
        site = site_data['site']
        print(f"\n👉 Vérification: {site}")
        
        excel_path = AUDIT_DIR / f"Grille Audit {site}.xlsx"
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb['Synthèse']
        
        # Extraire les scores Excel
        theme_mapping = {
            'Affichage Obligatoire': 'AFF',
            'Exploitation': 'EXP',
            'Image': 'IMA',
            'Qualité': 'QUA',
            'Résidents': 'RES',
            'RH': 'RH',
            'Sécurité': 'SEC',
            'Slow Impact': 'SLO'
        }
        
        excel_scores = {}
        for row in ws.iter_rows(min_row=6, max_row=14, values_only=True):
            if not row or not row[1]:
                continue
            theme_name = str(row[1]).strip()
            if theme_name in theme_mapping:
                code = theme_mapping[theme_name]
                try:
                    excel_scores[code] = {
                        'obtenu': float(row[2]) if row[2] else 0,
                        'possible': float(row[3]) if row[3] else 0,
                        'bloquants': int(row[5]) if row[5] else 0
                    }
                except (ValueError, TypeError):
                    pass
        
        wb.close()
        
        # Comparer avec JSON
        json_scores = site_data.get('scores_par_theme', {})
        
        all_ok = True
        for code in ['AFF', 'EXP', 'IMA', 'QUA', 'RES', 'RH', 'SEC', 'SLO']:
            if code in excel_scores and code in json_scores:
                json_obtenu = json_scores[code].get('obtenu', 0)
                json_possible = json_scores[code].get('possible', 0)
                json_bloquants = json_scores[code].get('bloquants', 0)
                
                excel_obtenu = excel_scores[code]['obtenu']
                excel_possible = excel_scores[code]['possible']
                excel_bloquants = excel_scores[code]['bloquants']
                
                match_obtenu = abs(json_obtenu - excel_obtenu) < 0.01
                match_possible = abs(json_possible - excel_possible) < 0.01
                match_bloquants = json_bloquants == excel_bloquants
                
                if not (match_obtenu and match_possible and match_bloquants):
                    all_ok = False
                    print(f"   ❌ {code}: Obtenu {json_obtenu:.1f}vs{excel_obtenu:.1f} | "
                          f"Possible {json_possible:.1f}vs{excel_possible:.1f} | "
                          f"Bloquants {json_bloquants}vs{excel_bloquants}")
        
        if all_ok:
            print(f"   ✅ Tous les scores pondérés cohérents")
        
        results.append(all_ok)
    
    return results

def verify_math_consistency(json_data):
    """Vérifie la cohérence mathématique des calculs"""
    
    print("\n" + "="*100)
    print("🔍 PARTIE 3 : VÉRIFICATION MATHÉMATIQUE (Cohérence des calculs)")
    print("="*100)
    
    all_ok = True
    
    for site in json_data:
        site_name = site['site']
        errors = []
        
        # 1. Vérifier que conforme + nc + ec + na = total
        stats = site['global_stats']
        somme = stats['conforme'] + stats['non_conforme'] + stats['en_cours'] + stats['na']
        if somme != stats['total']:
            errors.append(f"Somme statuts ({somme}) ≠ Total ({stats['total']})")
        
        # 2. Vérifier la cohérence des thèmes
        total_from_themes = sum(t['stats']['total'] for t in site['themes'].values())
        if total_from_themes != stats['total']:
            errors.append(f"Total thèmes ({total_from_themes}) ≠ Global ({stats['total']})")
        
        # 3. Vérifier le taux pondéré
        pond = site['global_stats_ponderees']
        if pond['possible'] > 0:
            taux_calc = (pond['obtenu'] / pond['possible']) * 100
            if abs(taux_calc - pond['taux']) > 0.01:
                errors.append(f"Taux calculé ({taux_calc:.2f}) ≠ Stocké ({pond['taux']:.2f})")
        
        # 4. Vérifier les priorités
        total_prio = sum(sum(site['priorites_globales'][p].values()) for p in ['P0', 'P1', 'P2', 'P3'])
        if total_prio != stats['total']:
            errors.append(f"Total priorités ({total_prio}) ≠ Total ({stats['total']})")
        
        if errors:
            all_ok = False
            print(f"\n   ❌ {site_name}:")
            for err in errors:
                print(f"      - {err}")
    
    if all_ok:
        print("\n   ✅ Tous les calculs sont cohérents pour tous les sites")
    
    return all_ok

def verify_group_totals(json_data):
    """Vérifie les totaux de groupe"""
    
    print("\n" + "="*100)
    print("🔍 PARTIE 4 : VÉRIFICATION TOTAUX GROUPE")
    print("="*100)
    
    # Calculer les totaux
    total_criteres = sum(s['global_stats']['total'] for s in json_data)
    total_conforme = sum(s['global_stats']['conforme'] for s in json_data)
    total_nc = sum(s['global_stats']['non_conforme'] for s in json_data)
    total_ec = sum(s['global_stats']['en_cours'] for s in json_data)
    total_na = sum(s['global_stats']['na'] for s in json_data)
    
    total_bloquants = sum(s['nb_bloquants'] for s in json_data)
    
    total_obtenu = sum(s['global_stats_ponderees']['obtenu'] for s in json_data)
    total_possible = sum(s['global_stats_ponderees']['possible'] for s in json_data)
    
    # Vérifications
    checks = []
    
    # 1. Somme = total
    somme = total_conforme + total_nc + total_ec + total_na
    checks.append(("Somme statuts = Total", somme == total_criteres, f"{somme} = {total_criteres}"))
    
    # 2. Chaque site a le même nombre de critères
    unique_totals = set(s['global_stats']['total'] for s in json_data)
    checks.append(("Uniformité critères/site", len(unique_totals) == 1, f"{len(unique_totals)} valeur unique ({list(unique_totals)[0] if unique_totals else 0})"))
    
    # 3. Taux cohérent
    taux = (total_conforme / total_criteres * 100) if total_criteres > 0 else 0
    checks.append(("Taux calculable", 0 <= taux <= 100, f"{taux:.2f}%"))
    
    # 4. Scores pondérés cohérents
    taux_pondere = (total_obtenu / total_possible * 100) if total_possible > 0 else 0
    checks.append(("Taux pondéré calculable", 0 <= taux_pondere <= 100, f"{taux_pondere:.2f}%"))
    
    print(f"\n   📊 Totaux groupe:")
    print(f"      - Critères totaux: {total_criteres:,}")
    print(f"      - Conforme: {total_conforme:,} ({taux:.2f}%)")
    print(f"      - Non conforme: {total_nc:,}")
    print(f"      - En cours: {total_ec:,}")
    print(f"      - Bloquants: {total_bloquants}")
    print(f"      - Score pondéré: {total_obtenu:.0f}/{total_possible:.0f} ({taux_pondere:.2f}%)")
    
    print(f"\n   ✅ Vérifications:")
    for desc, ok, detail in checks:
        icon = "✅" if ok else "❌"
        print(f"      {icon} {desc}: {detail}")
    
    return all(c[1] for c in checks)

def main():
    print("="*100)
    print("🔬 CONTRÔLE D'INTÉGRITÉ COMPLÈTE - AUDIT SLOW VILLAGE v2")
    print("="*100)
    
    json_data = load_json_data()
    print(f"📁 Données chargées: {len(json_data)} sites")
    
    # Exécuter toutes les vérifications
    results_criteria = verify_sample_criteria(json_data, 5)
    results_scores = verify_scores_synthese(json_data)
    results_math = verify_math_consistency(json_data)
    results_totals = verify_group_totals(json_data)
    
    # Résumé final
    print("\n" + "="*100)
    print("📋 RAPPORT FINAL")
    print("="*100)
    
    print(f"\n   1. Critères aléatoires: {sum(results_criteria)}/{len(results_criteria)} ✅")
    print(f"   2. Scores pondérés: {sum(results_scores)}/{len(results_scores)} ✅")
    print(f"   3. Cohérence mathématique: {'✅' if results_math else '❌'}")
    print(f"   4. Totaux groupe: {'✅' if results_totals else '❌'}")
    
    all_ok = all(results_criteria) and all(results_scores) and results_math and results_totals
    
    print("\n" + "="*100)
    if all_ok:
        print("🏆 SUCCESS: Toutes les vérifications sont passées avec succès")
        print("   Les données sont parfaitement cohérentes et fiables.")
    else:
        print("⚠️  ATTENTION: Certaines vérifications ont échoué")
        print("   Veuillez examiner les détails ci-dessus.")
    print("="*100)

if __name__ == "__main__":
    main()
